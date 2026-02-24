import os
import sys
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ===== ENTRADA (solo consola) =====
def obtener_argumentos():
    if len(sys.argv) < 2:
        raise ValueError(
            "Uso: python reporte.py <archivo.xlsx>"
        )

    archivo_entrada = sys.argv[1]
    archivo_salida = (
        sys.argv[2] if len(sys.argv) >= 3 else "reporte.xlsx"
    )

    fecha_inicio = (
        sys.argv[3] if len(sys.argv) >= 5 else None
    )
    fecha_fin = (
        sys.argv[4] if len(sys.argv) >= 5 else None
    )

    return archivo_entrada, archivo_salida, fecha_inicio, fecha_fin


# ===== VALIDACIÓN EXTERNA =====
def validar_archivo(archivo):
    if not os.path.exists(archivo):
        raise ValueError(f"El archivo {archivo} no existe.")


# ===== LECTURA =====
def leer_datos(archivo):
    try:
        return pd.read_excel(archivo, engine="openpyxl")
    except Exception as e:
        raise ValueError(f"No se pudo leer el archivo: {e}")


# ===== VALIDACIÓN INTERNA =====
def validar_columnas(df):
    columnas_requeridas = {
        "fecha",
        "vendedor",
        "producto",
        "cantidad",
        "precio"
    }

    faltantes = columnas_requeridas - set(df.columns)

    if faltantes:
        raise ValueError(
            f"Faltan columnas obligatorias: {faltantes}"
        )


# ===== FILTRO =====
def aplicar_filtro(df, fecha_inicio=None, fecha_fin=None):

    df["fecha"] = pd.to_datetime(
        df["fecha"],
        errors="coerce"
    )

    if df["fecha"].isna().any():
        raise ValueError(
            "Hay fechas inválidas en el archivo."
        )

    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio = pd.to_datetime(fecha_inicio)
            fecha_fin = pd.to_datetime(fecha_fin)
        except ValueError:
            raise ValueError(
                "Formato de fecha inválido. Usa YYYY-MM-DD"
            )

        df = df[
            (df["fecha"] >= fecha_inicio)
            & (df["fecha"] <= fecha_fin)
        ]

        if df.empty:
            raise ValueError(
                "No hay datos en el rango seleccionado."
            )

    return df


# ===== MÉTRICAS =====
def calcular_metricas(df):

    df["cantidad"] = pd.to_numeric(
        df["cantidad"],
        errors="coerce"
    )
    df["precio"] = pd.to_numeric(
        df["precio"],
        errors="coerce"
    )

    if df[["cantidad", "precio"]].isna().any().any():
        raise ValueError(
            "Hay valores no numéricos en cantidad o precio."
        )

    df["total"] = df["cantidad"] * df["precio"]

    ventas_por_dia = (
        df.groupby("fecha")["total"]
        .sum()
        .sort_index()
    )

    if ventas_por_dia.empty:
        raise ValueError(
            "No hay datos suficientes."
        )

    dia_mayor_venta = ventas_por_dia.idxmax()
    monto_dia_mayor_venta = ventas_por_dia.max()

    total_general = df["total"].sum()

    total_por_vendedor = (
        df.groupby("vendedor")["total"]
        .sum()
        .sort_values(ascending=False)
    )

    producto_mas_vendido = (
        df.groupby("producto")["cantidad"]
        .sum()
        .idxmax()
    )

    resumen = pd.DataFrame({
        "Métrica": [
            "Total General",
            "Producto Más Vendido",
            "Día con Mayor Venta",
            "Monto Día Mayor Venta"
        ],
        "Valor": [
            float(total_general),
            producto_mas_vendido,
            dia_mayor_venta.strftime("%Y-%m-%d"),
            float(monto_dia_mayor_venta)
        ]
    })

    return df, resumen, total_por_vendedor, ventas_por_dia


# ===== SALIDA =====
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def generar_reporte(
    archivo_salida,
    df,
    resumen,
    total_por_vendedor,
    ventas_por_dia
):
    try:
        with pd.ExcelWriter(
            archivo_salida,
            engine="openpyxl"
        ) as writer:

            df.to_excel(
                writer,
                sheet_name="Ventas Detalladas",
                index=False
            )

            resumen.to_excel(
                writer,
                sheet_name="Resumen",
                index=False
            )

            total_por_vendedor.to_excel(
                writer,
                sheet_name="Por Vendedor"
            )

            ventas_por_dia.to_excel(
                writer,
                sheet_name="Ventas por Día"
            )

            # ===== FORMATO =====
            workbook = writer.book

            for sheet in workbook.worksheets:

                # Encabezados en negrita
                for cell in sheet[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

                # Ajustar ancho columnas automáticamente
                for col in sheet.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)

                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(
                                    max_length,
                                    len(str(cell.value))
                                )
                        except:
                            pass

                    adjusted_width = max_length + 2
                    sheet.column_dimensions[
                        col_letter
                    ].width = adjusted_width

                # Congelar encabezado
                sheet.freeze_panes = "A2"

    except Exception as e:
        raise ValueError(
    "No hay ventas en el rango seleccionado."
)


# ===== MAIN (solo consola) =====
def main():
    archivo_entrada, archivo_salida, fi, ff = obtener_argumentos()

    validar_archivo(archivo_entrada)
    df = leer_datos(archivo_entrada)
    validar_columnas(df)
    df = aplicar_filtro(df, fi, ff)

    df, resumen, tpv, vpd = calcular_metricas(df)

    generar_reporte(archivo_salida, df, resumen, tpv, vpd)

    print(f"Reporte generado correctamente en {archivo_salida}")


if __name__ == "__main__":
    main()